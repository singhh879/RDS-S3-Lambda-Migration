import os
import getpass
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'user': 'marsquantMasterUser',
    'database': 'datafeeddatabase',
    'password': os.environ.get('DB_PASS') or getpass.getpass('DB Password: ')
}

BASE_TABLES = ['master_table', 'mcx_table', 'nifty50_table', 'upstox_table']

HEADER_FILL   = PatternFill('solid', start_color='1F4E79')
SUBHEAD_FILL  = PatternFill('solid', start_color='2E75B6')
ALT_FILL      = PatternFill('solid', start_color='EBF3FB')
GREEN_FILL    = PatternFill('solid', start_color='C6EFCE')
RED_FILL      = PatternFill('solid', start_color='FFC7CE')
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT     = Font(name='Arial', size=10)
BOLD_FONT     = Font(name='Arial', bold=True, size=10)
CENTER        = Alignment(horizontal='center', vertical='center')
LEFT          = Alignment(horizontal='left', vertical='center')

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def get_conn():
    return psycopg2.connect(**DB_CONFIG)

def get_all_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'datafeedschema' AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """)
    return [r[0] for r in cur.fetchall()]

def get_row_count(conn, table):
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) FROM datafeedschema.{table}")
    return cur.fetchone()[0]

def get_instrument_stats(conn, table):
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            instrument,
            COUNT(*)        AS total_rows,
            COUNT(delta)    AS delta_rows,
            COUNT(gamma)    AS gamma_rows,
            COUNT(theta)    AS theta_rows,
            COUNT(vega)     AS vega_rows,
            COUNT(iv)       AS iv_rows,
            COUNT(oi)       AS oi_rows,
            COUNT(up)       AS up_rows,
            MIN(tickd::text) AS data_from,
            MAX(tickd::text) AS data_to
        FROM datafeedschema.{table}
        GROUP BY instrument
        ORDER BY total_rows DESC
    """)
    cols = ['instrument', 'total_rows', 'delta_rows', 'gamma_rows', 'theta_rows',
            'vega_rows', 'iv_rows', 'oi_rows', 'up_rows', 'data_from', 'data_to']
    return pd.DataFrame(cur.fetchall(), columns=cols)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border()

def style_data_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = ALT_FILL if alt else PatternFill()
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = thin_border()

def write_summary_sheet(wb, conn, all_tables):
    ws = wb.active
    ws.title = 'Summary'

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'MarsQuant — Database Tables Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Schema: datafeedschema  |  Total Tables: {len(all_tables)}'
    ws['A2'].font = Font(name='Arial', size=10, color='595959')
    ws['A2'].alignment = CENTER
    ws.row_dimensions[2].height = 18

    headers = ['Table Name', 'Family', 'Row Count', 'Has Greeks', 'Has OI', 'Partition']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = h
    style_header_row(ws, 4, len(headers))
    ws.row_dimensions[4].height = 20

    print('Counting rows across all tables...')
    for i, table in enumerate(all_tables):
        print(f'  [{i+1}/{len(all_tables)}] {table}')
        row = i + 5
        count = get_row_count(conn, table)

        family = next((b for b in BASE_TABLES if table == b or table.startswith(b + '_')), 'other')
        is_partition = table not in BASE_TABLES
        has_greeks = 'Yes' if family in ['master_table', 'mcx_table', 'nifty50_table', 'upstox_table'] else 'Unknown'

        ws.cell(row=row, column=1).value = table
        ws.cell(row=row, column=2).value = family
        ws.cell(row=row, column=3).value = count
        ws.cell(row=row, column=4).value = has_greeks
        ws.cell(row=row, column=5).value = 'Yes'
        ws.cell(row=row, column=6).value = 'Yes' if is_partition else 'No (base)'

        style_data_row(ws, row, len(headers), alt=(i % 2 == 0))
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=3).number_format = '#,##0'

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.freeze_panes = 'A5'

def write_detail_sheet(wb, conn, table):
    print(f'\nQuerying instrument breakdown for {table}...')
    df = get_instrument_stats(conn, table)
    if df.empty:
        return

    ws = wb.create_sheet(title=table[:31])

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Greeks & Data Coverage — {table}'
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws['A1'].alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Total instruments: {len(df)}   |   Total rows: {df["total_rows"].sum():,}'
    ws['A2'].font = Font(name='Arial', size=10, color='595959')
    ws['A2'].alignment = CENTER

    headers = ['Instrument', 'Total Rows', 'Delta', 'Gamma', 'Theta',
               'Vega', 'IV', 'OI', 'Underlying (up)', 'Data From', 'Data To']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col).value = h
    style_header_row(ws, 4, len(headers))
    ws.row_dimensions[4].height = 20

    greek_cols = {'Delta': 3, 'Gamma': 4, 'Theta': 5, 'Vega': 6, 'IV': 7, 'OI': 8, 'Underlying': 9}

    for i, r in df.iterrows():
        row = i + 5
        total = r['total_rows']

        ws.cell(row=row, column=1).value  = r['instrument']
        ws.cell(row=row, column=2).value  = total
        ws.cell(row=row, column=3).value  = r['delta_rows']
        ws.cell(row=row, column=4).value  = r['gamma_rows']
        ws.cell(row=row, column=5).value  = r['theta_rows']
        ws.cell(row=row, column=6).value  = r['vega_rows']
        ws.cell(row=row, column=7).value  = r['iv_rows']
        ws.cell(row=row, column=8).value  = r['oi_rows']
        ws.cell(row=row, column=9).value  = r['up_rows']
        ws.cell(row=row, column=10).value = r['data_from']
        ws.cell(row=row, column=11).value = r['data_to']

        style_data_row(ws, row, len(headers), alt=(i % 2 == 0))

        # Highlight greeks cells green/red based on coverage
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            val = cell.value or 0
            cell.fill = GREEN_FILL if val > 0 else RED_FILL
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0'

        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=2).number_format = '#,##0'

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 14
    for col in range(3, 10):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.freeze_panes = 'A5'

def main():
    print('Connecting to database...')
    conn = get_conn()
    print('Connected.\n')

    wb = Workbook()
    all_tables = get_all_tables(conn)
    print(f'Found {len(all_tables)} tables in datafeedschema.\n')

    write_summary_sheet(wb, conn, all_tables)

    for table in BASE_TABLES:
        if table in all_tables:
            write_detail_sheet(wb, conn, table)

    output_path = os.path.expanduser('~/Downloads/MarsQuant/db_greeks_report.xlsx')
    wb.save(output_path)
    conn.close()
    print(f'\nReport saved to: {output_path}')

if __name__ == '__main__':
    main()
